import random
from openpyxl import load_workbook

FILE_PATH = "Auto_Filled_Bowling_Data.xlsx"
wb = load_workbook(FILE_PATH)

# ----------------------------
# Shot Maximum Runs Map
# ----------------------------
SHOT_MAX_RUNS = {
    "Cut": 4,
    "Late Cut": 4,
    "Square Cut": 4,
    "Pull": 6,
    "Hook": 6,
    "Straight Drive": 4,
    "Cover Drive": 4,
    "On Drive": 4,
    "Flick": 6,
    "Glance": 4,
    "Sweep": 6,
    "Reverse Sweep": 6,
    "Paddle Sweep": 6,
    "Lofted Straight": 6,
    "Lofted Cover": 6,
    "Lofted On": 6,
    "Upper Cut": 6,
    "Ramp Shot": 6,
    "Defense": 0,
    "Forward Defense": 0,
    "Backfoot Defense": 0,
    "Leave": 0
}


# ----------------------------
# Clamp Score
# ----------------------------
def clamp(value):
    return max(1, min(100, int(value)))


# ----------------------------
# Read Shot Rating from Excel
# ----------------------------
def get_shot_rating(
        bowling_type,
        line,
        length,
        variation,
        shot_type):

    wb = load_workbook(FILE_PATH)
    ws = wb[bowling_type]

    headers = [cell.value for cell in ws[1]]
    shot_col_index = headers.index(shot_type)

    for row in ws.iter_rows(min_row=2):
        if (row[0].value == line and
                row[1].value == length and
                row[2].value == variation):

            return row[shot_col_index].value

    raise ValueError("Combination not found in Excel")


# ----------------------------
# Calculate Effective Score
# ----------------------------
def calculate_effective_score(
        batsman_rating,
        bowler_rating,
        shot_rating):

    score = (
        0.8 * shot_rating +
        0.4 * batsman_rating -
        0.2 * bowler_rating
    )

    return clamp(score)



# ----------------------------
# Simulate Ball Outcome
# ----------------------------
def simulate_ball(
        batsman_rating,
        bowler_rating,
        bowling_type,
        line,
        length,
        variation,
        shot_type):

    shot_rating = get_shot_rating(
        bowling_type,
        line,
        length,
        variation,
        shot_type
    )

    effective_score = calculate_effective_score(
        batsman_rating,
        bowler_rating,
        shot_rating
    )

    max_runs = SHOT_MAX_RUNS.get(shot_type, 4)

    # Defense & Leave always dot
    if max_runs == 0:
        return {
            "Result": 0,
            "Type": "DOT",
            "Effective Score": effective_score
        }

    # Very High
    if effective_score >= 98:
        if max_runs == 6:
            return {"Result": 6, "Type": "SIX", "Effective Score": effective_score}
        else:
            return {"Result": 4, "Type": "FOUR", "Effective Score": effective_score}

    # High
    if 85 <= effective_score < 98:
        return {"Result": 4, "Type": "FOUR", "Effective Score": effective_score}

    # Medium
    if 75 <= effective_score < 85:
        return {"Result": 2, "Type": "TWO", "Effective Score": effective_score}

    # Low Case
    stump_lines = ["Off Stump", "Middle Stump", "Leg Stump"]

    if line in stump_lines and length in ["Yorker", "Good Length", "Full"]:
        wicket_roll = random.randint(1, 100)
        if wicket_roll > effective_score:
            return {"Result": "W", "Type": "WICKET", "Effective Score": effective_score}

    return {"Result": 0, "Type": "DOT", "Effective Score": effective_score}


def get_dropdown_data():
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        shots = headers[3:]

        lines = set()
        lengths = set()
        variations = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            lines.add(row[0])
            lengths.add(row[1])
            variations.add(row[2])

        data[sheet] = {
            "shots": shots,
            "lines": list(lines),
            "lengths": list(lengths),
            "variations": list(variations)
        }
    return data


# ----------------------------
# SIMPLE USER INPUT
# ----------------------------
if __name__ == "__main__":

    print("\n🏏 Ball Simulation Engine\n")

    batsman_rating = int(input("Enter Batsman Rating (1-100): "))
    bowler_rating = int(input("Enter Bowler Rating (1-100): "))

    bowling_type = input("Enter Bowling Type (Fast/Medium/Leg Spin/Off Spin): ")
    line = input("Enter Line: ")
    length = input("Enter Length: ")
    variation = input("Enter Variation: ")
    shot_type = input("Enter Shot Type: ")

    result = simulate_ball(
        batsman_rating,
        bowler_rating,
        bowling_type,
        line,
        length,
        variation,
        shot_type
    )

    print("\n--- Ball Result ---")
    print(result)
