import random
from openpyxl import load_workbook

FILE_PATH = "Auto_Filled_Bowling_Data.xlsx"

wb = load_workbook(FILE_PATH)

SHOT_MAX_RUNS = {
    "Cut": 4, "Late Cut": 4, "Square Cut": 4,
    "Pull": 6, "Hook": 6,
    "Straight Drive": 4, "Cover Drive": 4,
    "On Drive": 4, "Flick": 6, "Glance": 4,
    "Sweep": 6, "Reverse Sweep": 6,
    "Paddle Sweep": 6, "Lofted Straight": 6,
    "Lofted Cover": 6, "Lofted On": 6,
    "Upper Cut": 6, "Ramp Shot": 6,
    "Defense": 0, "Forward Defense": 0,
    "Backfoot Defense": 0, "Leave": 0
}

def clamp(value):
    return max(1, min(100, int(value)))

def get_shot_rating(bowling_type, line, length, variation, shot_type):
    ws = wb[bowling_type]
    headers = [cell.value for cell in ws[1]]
    shot_col_index = headers.index(shot_type)

    for row in ws.iter_rows(min_row=2):
        if (row[0].value == line and
            row[1].value == length and
            row[2].value == variation):
            return row[shot_col_index].value

    raise ValueError("Combination not found")

def calculate_effective_score(batsman_rating, bowler_rating, shot_rating):
    score = 0.8 * shot_rating + 0.4 * batsman_rating - 0.2 * bowler_rating
    return clamp(score)

def simulate_ball(batsman_rating, bowler_rating,
                  bowling_type, line, length, variation, shot_type):

    shot_rating = get_shot_rating(
        bowling_type, line, length, variation, shot_type)

    effective_score = calculate_effective_score(
        batsman_rating, bowler_rating, shot_rating)

    max_runs = SHOT_MAX_RUNS.get(shot_type, 4)

    if max_runs == 0:
        return {"Result": 0, "Type": "DOT"}

    if effective_score >= 98:
        return {"Result": max_runs, "Type": "BOUNDARY"}

    if 85 <= effective_score < 98:
        return {"Result": 4, "Type": "FOUR"}

    if 75 <= effective_score < 85:
        return {"Result": 2, "Type": "TWO"}

    stump_lines = ["Off Stump", "Middle Stump", "Leg Stump"]

    if line in stump_lines and length in ["Yorker", "Good Length", "Full"]:
        if random.randint(1, 100) > effective_score:
            return {"Result": "W", "Type": "WICKET"}

    return {"Result": 0, "Type": "DOT"}

def get_dropdown_data():
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        shots = headers[3:]

        lines = set()
        lengths = set()
        variations = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            lines.add(row[0])
            lengths.add(row[1])
            variations.add(row[2])

        data[sheet] = {
            "shots": shots,
            "lines": list(lines),
            "lengths": list(lengths),
            "variations": list(variations)
        }
    return data
